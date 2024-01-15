from tkinter import *
from tkinter import ttk
from tkinter import messagebox as mb
from tkinter.filedialog import asksaveasfilename
from imaplib import IMAP4,IMAP4_SSL
import email
from email.header import decode_header
from socket import gaierror
import os
import webbrowser
from openpyxl import Workbook

def all_mail():
    global screen1,tree_columns,tree

    def search_tree():
        if search_sbox.get() == "Date":
            if len(search_entry.get()) != 0:
                for child in tree.get_children():
                    if tree.item(child)["values"][1].casefold().find(
                        search_entry.get().casefold()) != -1:
                        tree.selection_add(child)

                    else:
                        pass
            else:
                pass

        elif search_sbox.get() == "Email":
            for child in tree.get_children():
                if len(search_entry.get()) != 0:
                    if tree.item(child)["values"][2].casefold().find(
                        search_entry.get().casefold()) != -1:
                        tree.selection_add(child)

                    else:
                        pass
                else:
                    pass

        elif search_sbox.get() == "Subject":
            for child in tree.get_children():
                if len(search_entry.get()) != 0:
                    if str(tree.item(child)["values"][3]).casefold().find(
                        search_entry.get().casefold()) != -1:
                        tree.selection_add(child)
                    
                    else:
                        pass
                else:
                    pass
        else:
            pass

    def save_sel(tv):
        if len(tv.selection()) != 0:
            if os.path.exists('./files'):
                file = asksaveasfilename(defaultextension = ".txt",
                initialdir = "./files/",
                filetypes = (("Text","*.txt"),
                ("CSV","*.csv"),
                ("Excel","*.xlsx")),
                title = "Save selection")

                if file :
                    if file.split(".")[-1] == "xlsx":
                        f = Workbook()
                        s = f.active
                        for i in tv.selection():
                            s.append([tv.item(i)["values"][0],
                            tv.item(i)["values"][1],
                            tv.item(i)["values"][2],
                            tv.item(i)["values"][3]])

                        f.save(file)

                    else:

                        with open(file,"w",encoding = 'utf-8') as f:
                            f.write(f"Sno,Date,Email,Subject\n")
                            for i in tv.selection():
                                if i != tv.selection()[-1]:
                                    f.write("{},{},{},{}\n".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                                else:
                                    f.write("{},{},{},{}".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                else:
                    pass

            else:
                os.mkdir('./files/')
                file = asksaveasfilename(defaultextension = ".txt",
                initialdir = "./files/",
                filetypes = (("Text","*.txt"),
                ("CSV","*.csv"),
                ("Excel","*.xlsx")),
                title = "Save selection")

                if file :

                    if file.split(".")[-1] == "xlsx":
                        f = Workbook()
                        s = f.active
                        for i in tv.selection():
                            s.append([tv.item(i)["values"][0],
                            tv.item(i)["values"][1],
                            tv.item(i)["values"][2],
                            tv.item(i)["values"][3]])

                        f.save(file)

                    else:

                        with open(file,"w",encoding = 'utf-8') as f:
                            f.write(f"Sno,Date,Email,Subject\n")
                            for i in tv.selection():
                                if i != tv.selection()[-1]:
                                    f.write("{},{},{},{}\n".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                                else:
                                    f.write("{},{},{},{}".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                else:
                    pass

        else:
            mb.showinfo("","Please select row(s) to save.")

    def save_all(tv):
        if len(tv.get_children()) != 0:
            if os.path.exists('./files/'):
                file = asksaveasfilename(defaultextension = ".txt",
                initialdir = "./files/",
                filetypes = (("Text","*.txt"),
                ("CSV","*.csv"),
                ("Excel","*.xlsx")),
                title = "Save all")

                if file :
                    if file.split(".")[-1] == "xlsx":
                        f = Workbook()
                        s = f.active
                        for i in tv.get_children():
                            s.append([tv.item(i)["values"][0],
                            tv.item(i)["values"][1],
                            tv.item(i)["values"][2],
                            tv.item(i)["values"][3]])

                        f.save(file)

                    else:
                        with open(file,"w",encoding = 'utf-8') as f:
                            f.write(f"Sno,Date,Email,Subject\n")
                            for i in tv.get_children():
                                if i != tv.get_children()[-1]:
                                    f.write("{},{},{},{}\n".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                                else:
                                    f.write("{},{},{},{}".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                else:
                    pass

            else:
                os.mkdir('./files')
                file = asksaveasfilename(defaultextension = ".txt",
                initialdir = "./files/",
                filetypes = (("Text","*.txt"),
                ("CSV","*.csv"),
                ("Excel","*.xlsx")),
                title = "Save all")

                if file :

                    if file.split(".")[-1] == "xlsx":
                        f = Workbook()
                        s = f.active
                        for i in tv.get_children():
                            s.append([tv.item(i)["values"][0],
                            tv.item(i)["values"][1],
                            tv.item(i)["values"][2],
                            tv.item(i)["values"][3]])

                        f.save(file)

                    else:
                        with open(file,"w",encoding = 'utf-8') as f:
                            f.write(f"Sno,Date,Email,Subject\n")
                            for i in tv.get_children():
                                if i != tv.get_children()[-1]:
                                    f.write("{},{},{},{}\n".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                                else:
                                    f.write("{},{},{},{}".format(
                                    tv.item(i)["values"][0],
                                    tv.item(i)["values"][1],
                                    tv.item(i)["values"][2],
                                    tv.item(i)["values"][3]))
                else:
                    pass

        else:
            mb.showinfo("","No row(s) available to save!")

    def delete(tv):
        if len(tv.selection()) == 1:
            tv.delete(tv.selection()) 

        elif len(tv.selection()) > 1:
            for i in tv.selection():
                tv.delete(i)

        else:
            mb.showinfo("","Please select row(s) to delete.")

    def reset():
        if len(tree.get_children()) != 0:
            for i in tree.get_children():
                    tree.delete(i)
        else:
            pass

        for i,j in enumerate(decoded_items,start = 1):
            tree.insert("",END,values = (i,j[0],j[1],j[2]))
        
    def tree_sort(tv,col,rev):
        l = [(tv.set(k, col), k) for k in tv.get_children()]
        l.sort(key = lambda v : v[0].casefold(),reverse = rev)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)
        tv.heading(col, text=col, command=lambda _col=col:tree_sort(tv, _col,not rev))

    def tree_reverse(tv,col,rev):
        l = [(tv.set(k, col), k) for k in tv.get_children()]
        l.sort(key = lambda v : int(v[0]),reverse = rev)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)
        tv.heading(col, text=col, command=lambda _col=col:tree_reverse(tv, _col,not rev))

    def refresh():
        if screen.call("ttk::style", "theme", "use") == "x-dark":
            if tree.selection():
                sel_rows_L.configure(text = str(len(tree.selection())))
                sel_rows_L.configure(image = sel_rows_img_dark)
            else:
                sel_rows_L.configure(text = str(len(tree.selection())))
                sel_rows_L.configure(image = sel_rows_img_dark_zero)
        else:
            if tree.selection():
                sel_rows_L.configure(text = str(len(tree.selection())))
                sel_rows_L.configure(image = sel_rows_img_light)
            else:
                sel_rows_L.configure(text = str(len(tree.selection())))
                sel_rows_L.configure(image = sel_rows_img_light_zero)
        
        tree.heading("#1",text = str(len(tree.get_children())))
        sel_rows_L.after(500,refresh) 

    screen1 = Toplevel(screen)
    screen1.geometry("1020x355")
    screen1.iconbitmap('./resource/Icon.ico')  
    screen1.title("All Mail")
    screen1.maxsize(height = 390,width = 1020)
    screen1.minsize(height = 355,width = 1020)

    frame = ttk.Frame(screen1)
    frame.pack(padx = 20,
    pady = (20,10))

    scrollbar_y = ttk.Scrollbar(frame)
    scrollbar_y.pack(side = RIGHT,fill = Y)

    scrollbar_x = ttk.Scrollbar(frame,orient = HORIZONTAL)
    scrollbar_x.pack(side = BOTTOM,fill = X)

    tree_columns = [str(len(decoded_items)),"Date","Email","Subject"]

    tree = ttk.Treeview(frame,
            height = 8,
            columns = tree_columns,
            yscrollcommand = scrollbar_y.set,
            xscrollcommand = scrollbar_x.set)

    scrollbar_y.config(command = tree.yview)
    scrollbar_x.config(command = tree.xview)

    tree.pack()

    tree.column("#0",width = 0,stretch = False)
    tree.column("#1",width = 100,minwidth = 120,anchor = CENTER,stretch = False)
    tree.column("#2",width = 110,minwidth = 150,anchor = CENTER,stretch = False)
    tree.column("#3",width = 280,minwidth = 320,anchor = W,stretch = False)
    tree.column("#4",width = 400,minwidth = 550,anchor = W)

    tree.heading("#1",text = str(len(decoded_items)),
    command = lambda _col = tree_columns[0]:tree_reverse(tree, _col,False))
    tree.heading("#2",text = "Date")
    tree.heading("#3",text = "Email",
    command = lambda _col = tree_columns[2]:tree_sort(tree, _col,False))
    tree.heading("#4",text = "Subject",
    command = lambda _col = tree_columns[3]:tree_sort(tree, _col,False))
    
    options = ["Subject","Email","Date"]

    search_sbox = ttk.Spinbox(screen1,
        values = options,
        state = "readonly",
        width = 8)
    search_sbox.pack(side = "left",pady = (0,20),padx = (40,25))

    search_entry = ttk.Entry(screen1,
        width = 20)     
    search_entry.pack(side = "left",pady = (0,20),padx = (0,25))

    search_B = ttk.Button(screen1,
        text = "Search",
        command = search_tree,
        width = 8)
    search_B.pack(side = "left",pady = (0,20),padx = (0,25))

    sel_rows_img_dark_zero = PhotoImage(file = "./resource/dark-lab-zero.gif")
    sel_rows_img_dark = PhotoImage(file = "./resource/dark-lab.gif")

    sel_rows_img_light_zero = PhotoImage(file = "./resource/light-lab-zero.gif")
    sel_rows_img_light = PhotoImage(file = "./resource/light-lab.gif")

    sel_rows_L = Label(screen1,
                text = str(len(tree.selection())),
                foreground = "#ffffff",
                font = ("Segoe UI",9),
                image = sel_rows_img_dark_zero,
                compound = "center",
                borderwidth = 0,
                relief = FLAT)
    sel_rows_L.pack(side = "left",pady = (0,20),padx = (0,25))
    
    save_all_B = ttk.Button(screen1,
        text = "Save all",
        command = lambda tv = tree:save_all(tv),
        width = 8)
    save_all_B.pack(side = "left",pady = (0,20),padx = (0,25))

    save_sel_B = ttk.Button(screen1,
        text = "Save sel",
        command = lambda tv = tree:save_sel(tv),
        width = 8)
    save_sel_B.pack(side = "left",pady = (0,20),padx = (0,25))

    delete_B = ttk.Button(screen1,
        text = "Delete",
        command = lambda tv = tree:delete(tv),
        width = 8)
    delete_B.pack(side = "left",pady = (0,20),padx = (0,25))

    reset_B = ttk.Button(screen1,
        text = "Reset",
        command = reset,
        width = 8)
    reset_B.pack(side = "left",pady = (0,20),padx = (0,30))

    for i,j in enumerate(decoded_items, start = 1):
        tree.insert("",END,values = (i,j[0],j[1],j[2]))

    refresh()
    screen1.mainloop()

def login():
    global screen,login_check,mail,decoded_items,logout_B
    def date(line):
        if len(line.split()) == 6:
            return " ".join(line.split()[1:-2])
        elif len(line.split()) == 7:
            return " ".join(line.split()[1:-3])

    def subject(line):
        subject_list = decode_header(line)
        sub_list = []
        for subject in subject_list:
            if subject[1]:
                subject = (subject[0].decode(subject[1]))
            elif type(subject[0]) == bytes:
                subject = subject[0].decode('utf-8')
            else:
                subject = subject[0]
            sub_list.append(subject)
        subject = ''.join(sub_list)
        subject = " ".join(subject.splitlines())
        return subject

    try:
        mail = IMAP4_SSL("imap.gmail.com")
        if mail:
            mail_id_entry = mail_id.get()
            pswrd_entry = pswrd.get()

            if not mail_id_entry and not pswrd_entry:
                mail_id.state(["invalid"])
                pswrd.state(["invalid"])
                mb.showinfo("","Enter your credentials!")

            elif not mail_id_entry :
                mail_id.state(["invalid"])
                mb.showinfo("","Enter the email address!")

            elif not pswrd_entry:
                pswrd.state(["invalid"])
                mb.showinfo("","Enter the password!")

            elif mail_id_entry and pswrd_entry: 
                login_check,_ = mail.login(mail_id_entry.strip(),pswrd_entry.strip())
                
                if login_check == "OK":
                    mail_id.state(["focus"])
                    pswrd.state(["focus"])
                    mail.select('"INBOX"')

                    data_check,data = mail.uid("search",None,"ALL")
                    inbox_items = data[0].split()
                    decoded_items = []

                    for item in inbox_items:
                        _,raw_email_data = mail.uid("fetch",item,'(RFC822)')
                        email_ = email.message_from_bytes(raw_email_data[0][1])
                        decoded_items.append([date(email_["date"]),
                        email_["From"].split()[-1][1:-1],
                        subject(email_['subject'])])
        
                else:
                    pass
            else:
                pass        
        else:
            pass
     
    except IMAP4.error:
        mail_id.state(["invalid"])
        pswrd.state(["invalid"])
        mb.showinfo("","Invalid credentials!")

    except gaierror:
        mb.showinfo("","Check your network connection!")

    finally:
        if data_check == "OK":
            login_B.place(x = 90, y = 88)

            logout_B = ttk.Button(screen,
                    text = "Logout",
                    command = logout,
                    width = 9)
            logout_B.place(x = 195, y = 88)
            logout_B.config(style = "Accent.TButton")
            all_mail()

def logout():
    global screen,login_check,mail,mail_id,pswrd
    try:
        if login_check == "OK":
            response = mb.askyesno("","Do you want to logout?")

            if response == 1:
                logout_check,_ = mail.logout()

                if logout_check == "BYE":
                    mail_id.delete(0,"end")
                    pswrd.delete(0,"end")
                    logout_B.destroy()
                    login_B.place(x = 140, y = 88)

                    if screen1.winfo_exists():
                        if mb.askyesno("","Do you want to close the All Mail window?"):
                            screen1.destroy()
                        else:
                            pass
                    else:
                        pass
                else:
                    pass
            else:
                pass
        else:
            pass
    except:
        pass

def theme_change():
    if screen.call("ttk::style", "theme", "use") == "x-dark":
        screen.call("set_theme", "light")
        theme_B["image"] = dark_img
        theme_B["activebackground"] = "#fafafa"

        if pswrd.cget('show') == '*':
            eye_B['image'] = black_show_img
        else:
            eye_B['image'] = black_hide_img
        eye_B["activebackground"] = "#fafafa"

        if info_screen.winfo_exists():
            link_1_B['activebackground'] = "#fafafa"
            link_2_B['activebackground'] = "#fafafa"

        else:
            pass

    else:
        screen.call("set_theme", "dark")
        theme_B['image'] = light_img
        theme_B['activebackground'] = "#1c1c1c"

        if pswrd.cget('show') == '*':
            eye_B['image'] = white_show_img
        else:
            eye_B['image'] = white_hide_img
        eye_B["activebackground"] = "#1c1c1c"
        

        if info_screen.winfo_exists():
            link_1_B['activebackground'] = "#1c1c1c"
            link_2_B['activebackground'] = "#1c1c1c"

        else:
            pass

def info_screen():
    global info_screen,link_1_B,link_2_B

    def click(url):
        webbrowser.open_new_tab(url)

    info_screen = Toplevel(screen)
    info_screen.title("Info")
    info_screen.iconbitmap("./resource/Icon.ico")
    info_screen.geometry("555x550")
    info_screen.resizable(0,0)

    info_frame = ttk.Frame(info_screen)
    info_frame.pack(fill = "both")     

    label_f_1 = ttk.LabelFrame(info_frame,
                text = "IMAP access",
                padding = (10,7))

    label_f_1.pack(fill = "both",
                expand = "no",
                pady = 10,
                padx = 20)

    text_1 = "1) On your computer, open https://mail.google.com/ \
    \n2) In the top right, click Settings Settings and then See all  settings.\
    \n3) Click the Forwarding and POP/IMAP tab.\
    \n4) In the 'IMAP access' section, select Enable IMAP.\
    \n5) Click Save Changes."

    label_1 = ttk.Label(label_f_1,
                font = ("Segoe UI",10,"bold"),
                text = text_1)

    label_1.pack(fill = "x")

    link_img = PhotoImage(file = "./resource/e_link.png")

    link_1_B = Button(info_frame,
                image = link_img,
                borderwidth = 0,
                relief = FLAT,
                activebackground = "#1c1c1c",
                command= lambda url ="https://mail.google.com/":click(url))
    link_1_B.place(x = 360, y = 36)

    label_f_2 = ttk.LabelFrame(info_frame,
                #font = ("Segoe UI",10,"bold"),
                text = "2-Step Verification",
                padding = (10,7))

    label_f_2.pack(fill = "both",
                expand = "no",
                pady = (0,10),
                padx = 20)

    text_2 = '1) Open your Google Account.\
    \n2) In the navigation panel, select Security.\
    \n3) Under “Signing in to Google” select 2-Step Verification and then Get started.\
    \n4) Follow the on-screen steps.'

    label_2 = ttk.Label(label_f_2,
                font = ("Segoe UI",10,"bold"),
                text = text_2)

    label_2.pack(fill = "x")

    link_2_B = Button(info_frame,
                image = link_img,
                borderwidth = 0,
                relief = FLAT,
                activebackground = "#1c1c1c",
                command= lambda url = "https://myaccount.google.com/":click(url))

    link_2_B.place(x = 217, y = 160)


    label_f_3 = ttk.LabelFrame(info_frame,
                text = "App Password",
                padding = (10,7))

    label_f_3.pack(fill = "both",
                expand = "no",
                pady = (0,20),
                padx = 20)

    text_3 = '1) Now under "Signing in to Google" select App Passwords.\
    \nIf you do not have this option, it might be because:\
    \n  a. 2-Step Verification is not set up for your account.\
    \n  b. 2-Step Verification is only set up for security keys.\
    \n  c. Your account is through work, school, or other organization.\
    \n  d. You turned on Advanced Protection.\
    \n2) At the bottom, choose Select app and choose the app you are using > \
    \n Select device and choose the device you are using > Generate.\
    \n3) Follow the instructions to enter the App Password.\
    \nThe App Password is the 16-character code in the yellow bar on your device.\
    \nNOTE: Use this password to login(Note down the password for future use)\
    \n4)Tap Done.'


    label_3 = ttk.Label(label_f_3,
                font = ("Segoe UI",10,"bold"),
                text = text_3)

    label_3.pack(fill = "x")


    done_B = ttk.Button(info_frame,
                text = "Done",
                command = lambda:info_screen.destroy(),
                width = 8,
                style = "Accent.TButton")

    done_B.pack(anchor = E,
                side = BOTTOM,
                padx = (0,25),
                pady = (0,16))

    info_screen.mainloop()
   
def main_screen():
    global screen,pswrd,mail_id,login_check,theme_B,dark_img,light_img,login_B,eye_B,white_show_img,white_hide_img,black_show_img,black_hide_img

    def show():
        if screen.call("ttk::style", "theme", "use") == "x-dark":
            eye_B["activebackground"] = "#1c1c1c"
            if pswrd.cget('show') == '*':
                pswrd.config(show = "")
                eye_B.config(image = white_hide_img)

            else:
                pswrd.config(show = "*")
                eye_B.config(image = white_show_img)
        else:
            eye_B["activebackground"] = "#fafafa"
            if pswrd.cget('show') == '*':
                pswrd.config(show = "")
                eye_B.config(image = black_hide_img)

            else:
                pswrd.config(show = "*")
                eye_B.config(image = black_show_img)


    screen = Tk()
    screen.call("source", "./resource/x.tcl")
    screen.call("set_theme", "dark")
    screen.title("Login")
    screen.iconbitmap('./resource/Icon.ico')
    screen.geometry("340x130")
    screen.resizable(0,0)

    ttk.Label(screen,
        text = "  Email",
        font = ("Segoe UI",10,"bold")).place(x = 27, y = 18)

    ttk.Label(screen,
        text = "Password",
        font = ("Segoe UI",10,"bold")).place(x = 10, y = 55)

    mail_id = ttk.Entry(screen,
            width = 30)
    mail_id.place(x = 80, y = 10)

    pswrd = ttk.Entry(screen,
            width = 30)
    pswrd.place(x = 80, y = 47)
    pswrd.config(show = "*")

    dark_img = PhotoImage(file = "./resource/dark-m.png")
    light_img = PhotoImage(file = "./resource/light-m.png")

    theme_B = Button(screen,
                image = light_img,
                borderwidth = 0,
                relief = SUNKEN,
                activebackground = "#1c1c1c",
                command = theme_change)
    theme_B.place(x = 12, y = 95)

    black_show_img = PhotoImage(file = "./resource/black_show.png")
    black_hide_img = PhotoImage(file = "./resource/black_hide.png")
    white_show_img = PhotoImage(file = "./resource/white_show.png")
    white_hide_img = PhotoImage(file = "./resource/white_hide.png")


    eye_B = Button(screen,
                image = white_show_img,
                borderwidth = 0,
                relief = SUNKEN,
                activebackground = "#1c1c1c",
                command = show)
    eye_B.place(x = 314,y = 55)

    login_B = ttk.Button(screen,
            text = "Login",
            command = login,
            width = 9)
    login_B.place(x = 140, y = 88)


    screen.after(300,info_screen())
    screen.mainloop()

main_screen()